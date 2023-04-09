import openpyxl


class HomePageData:

    #test_homepage_data = [{"name":"test1 name", "email":"test1@gmail.com", "password":"test1pswd", "gender":"Female", "dob":"12/31/2000", "status":"employed"}, {"name":"test2 name", "email":"test2@gmail.com", "password":"test2pswd", "gender":"Male", "dob":"06/14/1999", "status":"student"}]
    @staticmethod
    def getTestData(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook("C:\\Users\\srika\\PycharmProjects\\PythonSeleniumFramework\\testData\\signup_info.xlsx")
        #sheet = book.active
        sheet = book["Sheet1"]
        for i in range(1, sheet.max_row + 1):  # to get rows
            if sheet.cell(row=i, column=1).value == test_case_name:

                for j in range(2, sheet.max_column + 1):  # to get columns
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return [Dict]
